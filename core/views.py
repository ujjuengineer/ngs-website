from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .models import ContactMessage, CompanyCertificate, DailyReport, PDFRecord, IndexingRecord, UploadingRecord, QCRecord, MetadataRecord
from django.urls import reverse_lazy, reverse
from django.views.generic.edit import CreateView, UpdateView
from .forms import CertificateForm, DailyReportForm, DailyReportBulkForm, DailyReportUpdateForm
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic import DetailView
import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from django.forms import modelformset_factory
from django.db.models import Sum, Q
from django.core.paginator import Paginator
from datetime import date, timedelta
from django.utils import timezone
from .utils import sync_workflow_records
import datetime
from itertools import chain


# ── PUBLIC PAGES ──
def home(request):
    return render(request, 'core/home.html', {
        'active_nav': 'home',
        'process_steps': [
            ('01', 'Collection'),
            ('02', 'Scanning'),
            ('03', 'PDF & OCR'),
            ('04', 'Metadata'),
            ('05', 'Portal Upload'),
        ],
        'gallery': [
            {
                'src': 'https://lh3.googleusercontent.com/aida-public/AB6AXuAdT2EXoQ4W2O0UHQkdTsajaLUUamR8TPgzyNxweGJHftYg7iMB1BcMPIrHWG2xb-4yUXGhdKPeu661zXpheVJImAtROcQ-WRfr2JjflaD3ahUYkCwtbD50DOqPcqgFTilu9G253SQTsWUuF1gnJTrG2Ml9TtbxCtUsJSve0_U0VaSD58gAuU6eA0EHwlPvE-71_so1Myyd8oFu852Pf-wvYhWqH36zvYJVlAbIxRzQY2bxRMjnsL0GgmWBfgTDl749FBZhTHpNb4k',
                'alt': 'Institutional digitization workspace',
                'tag': 'Operations',
                'title': 'Secure document floors',
                'span': 'bento-lg',
                'tall': True,
            },
            {
                'src': 'https://lh3.googleusercontent.com/aida-public/AB6AXuDtPDI0wFjeuB5iva7FgWkY84zfgUhdPOs-k9ZqUy_PwlO6fDEVFGKypnyqTc4lfqlsFGIfyXWVbc4aj9uSL3WSCtO9lpc08euk5ynfy75BhkBdBrQz9uOpTiZUM3tpdj1JGsBo-h1uIzanTvY2pKRC0g6mTfd7a4IPlp9OC1XaONLIQM31VdIGOWS0ni9BdV8wnBwsTWZ404oKwmNiEhhadnJP1Pcx6z0MaqnhszpiIMBGwBULdrd2wjgmRMzm1yTr0FzbsH8XU6c',
                'alt': 'High-fidelity scanning equipment',
                'tag': 'Scanning',
                'title': 'Archival-grade capture',
                'span': 'bento-md',
                'tall': True,
            },
            {
                'src': 'https://lh3.googleusercontent.com/aida-public/AB6AXuCKOSR77azr8Pf6mVDU6Sn_khr9DW9gxXjAzH5goke5rf8oSllw5yBWF0da8q2-MCh_-62fUd_ypgK7sYC1fTGMJJ24szQSBGfw67n62kXk_cyoPfQ7Hsg4CS-h3fPLpQqcL6q0VOgS2qkDUHEKxkWb4IFct8tdgsb7IvoWqBgipoXQfex4IMMTg-W4BM234LhcoX5zv6ot-XTYle-uepGmDYD9ZCHfdp7FtEyIK_tgxfa27PyCluqgRDGnR8BB6rosi9y1oqyQFF4',
                'alt': 'Trained operations staff',
                'tag': 'People',
                'title': 'Skilled field teams',
                'span': 'bento-third',
                'tall': False,
            },
            {
                'src': 'https://lh3.googleusercontent.com/aida-public/AB6AXuCxpaeJCIqf6DKV8_e7lmmXLpU8M-4o4Begd-ffzsdBBgsHxJfML2HdWKrqmEVhsVMhBmnf40RCvnTbGRaFAmtBtgG9g26MZ9Tyg2e7rnfDgczsCHCkGGcu1wvTOPCRJIS0srBOKiZ8yVdDibHBtkTW28GcbjlA3YoYzdC9oOfO2c_NlcnR4Ner56ZQ-xPCFYXBgIm6p5Y0RzoFP2hANGk0U6YhtMI7vf91QoI9X5RmLprxiIVpCtuRDwlFs2JlCKRbGLCeNPrJrfY',
                'alt': 'IT infrastructure',
                'tag': 'Systems',
                'title': 'Reliable IT backbone',
                'span': 'bento-third',
                'tall': False,
            },
            {
                'src': 'https://lh3.googleusercontent.com/aida-public/AB6AXuBRUi61UufxyOwAE6SI2mVeA_Dccf1GbbER8XHMWiRFEYEuHiQan3XLkROwXRRA3vWCCns2dW6ZeQnFR3ndQlODJFvg3VbQ74U18nRE1MeDcamlz-H5lkW9wBR4YnAlJFrf8nXk4jZK0s7lyxwrR7SiyQ5uT0IPA4jg7ML82ZVj070RJ5C3DTxnHwm-JYy95BVhVxHgCIrxi2rA49K-4xss0Ti1kG8CMjxlzPbGP9d3EwT4uqMRYzL-OG-8rIPqkk1iy1r5yrT97Eo',
                'alt': 'Portal upload dashboard',
                'tag': 'Delivery',
                'title': 'Portal-ready outputs',
                'span': 'bento-third',
                'tall': False,
            },
            {
                'src': 'https://lh3.googleusercontent.com/aida-public/AB6AXuDfeDmruG-ctQKAlcfOaE-vK4x73qGnBCOJZu3NMSdOye2unDPXnpiONZmvNzn2sdSkb1iK79yb5fQlhnjujLWzxaf7xnOZxd4npTOS7B9XCjjjfdRFEKA8bAVk_1UKvsNkm_tSccGvJYNWxmLf-5k2V7uM6qyEbEZjGPsO-3tztbtvP8iMVitIXN_VUy0Cm4zck62Lhy4-SMrsE_YAtNCV14MCjvHnk4K_UDWZmgkiejqwLj5VtqxnRm-eNlK1j2seOpi4YEdV5G4',
                'alt': 'Document collection and storage',
                'tag': 'Custody',
                'title': 'Chain-of-custody intake',
                'span': 'bento-half',
                'tall': False,
            },
            {
                'src': 'https://lh3.googleusercontent.com/aida-public/AB6AXuCDWqhgejIs1igG6X6QGBZDOqxS5YMOqrEswnSfKLSSrtKU-vZ3t6bd_5nSkUcCVqLz6gzyivVogAA8t_95grAaI5TM1CKy4sEQBKOY35hkeAukpH1lcv-YiwK9tvAsqlEDsy5jtY9J849rseBf2fY_dkoFLetJ7cILbGTpwirQF-2jaSEF_PbNHIpvnCwXgCGwIu1EQ4UVl0elflLvUwKPggN-nKCepfIprPwHDkbHuDFWYYDeyF3GZOAivY9l3rtfPfYD8SOmhkg',
                'alt': 'Careful document handling',
                'tag': 'Quality',
                'title': 'Precision handling',
                'span': 'bento-half',
                'tall': False,
            },
        ],
    })


def about(request):
    return render(request, 'core/about.html', {
        'active_nav': 'about',
        'trust_items': [
            ('account_balance', 'Revenue Dept'),
            ('gavel', 'Judicial Records'),
            ('local_hospital', 'Health Archives'),
            ('school', 'Education Board'),
        ],
    })


def services(request):
    return render(request, 'core/services.html', {'active_nav': 'services'})


def process(request):
    return render(request, 'core/process.html', {
        'active_nav': 'process',
        'steps': [
            {
                'num': '01',
                'title': 'Document Collection',
                'desc': 'Secure intake of physical materials. Every batch is logged with a unique identifier for chain-of-custody tracking.',
                'icon': 'inventory_2',
                'image': 'https://lh3.googleusercontent.com/aida-public/AB6AXuDfeDmruG-ctQKAlcfOaE-vK4x73qGnBCOJZu3NMSdOye2unDPXnpiONZmvNzn2sdSkb1iK79yb5fQlhnjujLWzxaf7xnOZxd4npTOS7B9XCjjjfdRFEKA8bAVk_1UKvsNkm_tSccGvJYNWxmLf-5k2V7uM6qyEbEZjGPsO-3tztbtvP8iMVitIXN_VUy0Cm4zck62Lhy4-SMrsE_YAtNCV14MCjvHnk4K_UDWZmgkiejqwLj5VtqxnRm-eNlK1j2seOpi4YEdV5G4',
            },
            {
                'num': '02',
                'title': 'High-Fidelity Scanning',
                'desc': 'Industrial-grade optical equipment captures documents at archival resolution. Fragile materials use non-destructive flatbed scanners.',
                'icon': 'document_scanner',
                'image': 'https://lh3.googleusercontent.com/aida-public/AB6AXuCDWqhgejIs1igG6X6QGBZDOqxS5YMOqrEswnSfKLSSrtKU-vZ3t6bd_5nSkUcCVqLz6gzyivVogAA8t_95grAaI5TM1CKy4sEQBKOY35hkeAukpH1lcv-YiwK9tvAsqlEDsy5jtY9J849rseBf2fY_dkoFLetJ7cILbGTpwirQF-2jaSEF_PbNHIpvnCwXgCGwIu1EQ4UVl0elflLvUwKPggN-nKCepfIprPwHDkbHuDFWYYDeyF3GZOAivY9l3rtfPfYD8SOmhkg',
            },
            {
                'num': '03',
                'title': 'PDF Creation & OCR',
                'desc': 'Images convert to PDF/A for long-term preservation. OCR makes text fully searchable.',
                'icon': 'picture_as_pdf',
                'image': None,
                'badge': 'OCR Processing',
            },
            {
                'num': '04',
                'title': 'Metadata Extraction',
                'desc': 'Key data points are extracted and validated so documents can be accurately indexed and retrieved.',
                'icon': 'data_object',
                'image': 'https://lh3.googleusercontent.com/aida-public/AB6AXuC3BxVwZ0Kokdbwi7l0i3p-Wg7qsCwVLT58qvG-oZ3lomnAGeGKufmPY1s50pfCYiG8ul03seb1KtZSQFoc6MSEPshdEbkNjxnx0I20U9y26Z9ekkhYhsXVu2yAM28jPheOw54enI2emW8l7gTnmxEA1EXgAkhqW5L6Qh9MeXje-8CXnomUyBYvdPAhhbJOU52Y2i2YpkLv7dX7lnX30GL0_WzVgg8nJFy3fduF2UrL4C_d_17zRYm0PuerlXoBhtzuY9OuVBG2sjc',
            },
            {
                'num': '05',
                'title': 'Secure Portal Upload',
                'desc': 'Finalized assets and metadata transfer securely to the designated government portal or DMS.',
                'icon': 'cloud_upload',
                'image': 'https://lh3.googleusercontent.com/aida-public/AB6AXuBRUi61UufxyOwAE6SI2mVeA_Dccf1GbbER8XHMWiRFEYEuHiQan3XLkROwXRRA3vWCCns2dW6ZeQnFR3ndQlODJFvg3VbQ74U18nRE1MeDcamlz-H5lkW9wBR4YnAlJFrf8nXk4jZK0s7lyxwrR7SiyQ5uT0IPA4jg7ML82ZVj070RJ5C3DTxnHwm-JYy95BVhVxHgCIrxi2rA49K-4xss0Ti1kG8CMjxlzPbGP9d3EwT4uqMRYzL-OG-8rIPqkk1iy1r5yrT97Eo',
            },
        ],
    })


def contact(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()

        if name and email and subject and message:
            ContactMessage.objects.create(
                name=name, email=email, phone=phone,
                subject=subject, message=message
            )
            messages.success(request, 'Thank you! Your message has been received. We will contact you shortly.')
        else:
            messages.error(request, 'Please fill in all required fields.')
        return redirect('contact')
    return render(request, 'core/contact.html', {'active_nav': 'contact'})


# ── AUTH: LOGIN ──
def login_view(request):
    # Already logged in → send admin to admin panel, others home
    if request.user.is_authenticated:
        if request.user.is_staff or request.user.is_superuser:
            return redirect('/admin/')
        return redirect('home')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            display_name = user.get_full_name().strip() or user.username

            # Prefer an explicit safe next URL when present
            next_url = request.GET.get('next', '').strip()
            if next_url.startswith('/') and not next_url.startswith('//'):
                messages.success(request, f'Welcome back, {display_name}! Login successful.')
                return redirect(next_url)

            # Admin / staff with admin access → Django admin
            if user.is_staff or user.is_superuser:
                messages.success(request, f'Welcome, {display_name}! Opening admin panel.')
                return redirect('/admin/')

            # Regular staff / scanners → public site with portal nav
            messages.success(request, f'Welcome back, {display_name}! Login successful.')
            return redirect('home')

        messages.error(request, 'Invalid username or password. Please try again.')

    return render(request, 'core/login.html', {
        'active_nav': 'login',
        'hide_site_nav': True,
    })


# ── AUTH: LOGOUT ──
@login_required
def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('home')


# ── CERTIFICATES ──
class CertificateCreateView(CreateView):
    model = CompanyCertificate
    form_class = CertificateForm
    template_name = 'core/add_certificate.html'
    success_url = reverse_lazy('add_certificate')
    success_message = "Certificate for %(recipient_name)s was created successfully!"


class CertificateVerifyView(DetailView):
    model = CompanyCertificate
    template_name = 'core/verify_certificate.html'
    context_object_name = 'certificate'
    slug_field = 'certificate_number'
    slug_url_kwarg = 'certificate_number'

from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages

@login_required
def daily_report_create_view(request):
    # 1. Determine the auto-name for the logged-in user
    user = request.user
    full_name = user.get_full_name().strip()
    auto_name = full_name if full_name else user.username

    duplicate_report = None

    # 2. Handle Form Submission (POST request)
    if request.method == 'POST':
        form = DailyReportForm(request.POST)
        if form.is_valid():
            # Extract and strip spaces/periods from front and back
            raw_year = form.cleaned_data.get('year') or ''
            raw_volume = form.cleaned_data.get('volume_num') or ''
            
            cleaned_year = str(raw_year).strip(" .")
            cleaned_volume = str(raw_volume).strip(" .")
            location = form.cleaned_data.get('location')

            # Check if an identical report already exists using __iexact or direct matching
            # Note: This assumes your DB values might also need standardizing, 
            # or we match against the newly cleaned formats.
            existing = DailyReport.objects.filter(
                year__iexact=cleaned_year, 
                volume_num__iexact=cleaned_volume, 
                location=location
            ).first()

            if existing:
                duplicate_report = {
                    'pk': existing.pk,
                    'year': existing.year,
                    'volume_num': existing.volume_num,
                    'location': existing.get_location_display(),
                    'name': existing.name,
                    'created_at': existing.created_at,
                }

                return render(request, 'core/add_report.html', {
                    'form': form, # Keeps their input data on screen
                    'next': request.GET.get('next', ''),
                    'auto_name': auto_name,
                    'duplicate_report': duplicate_report,
                })

            # Save the new report with the auto_name enforced
            instance = form.save(commit=False)
            instance.name = auto_name
            
            # Apply the stripped values to the instance before saving
            instance.year = cleaned_year
            instance.volume_num = cleaned_volume
            
            instance.save()

            # Create the submodel
            sync_workflow_records(instance, request.user)

            # Success notification
            messages.success(
                request, 
                f"Daily report for {instance.name} at {instance.get_location_display()} recorded successfully!"
            )

            # Determine where to redirect next
            next_url = request.GET.get('next') or request.POST.get('next')
            if next_url:
                return redirect(next_url)
            return redirect(reverse('add_daily_report'))

    # 3. Handle Initial Page Load (GET request)
    else:
        initial_data = {'name': auto_name}
        for field in ('year', 'volume_num', 'location'):
            value = request.GET.get(field, '').strip()
            if value:
                initial_data[field] = value
        form = DailyReportForm(initial=initial_data)

    context = {
        'form': form,
        'next': request.GET.get('next', ''),
        'auto_name': auto_name,
    }
    return render(request, 'core/add_report.html', context)

# @login_required
# def daily_report_create_view(request):
#     # 1. Determine the auto-name for the logged-in user
#     user = request.user
#     full_name = user.get_full_name().strip()
#     auto_name = full_name if full_name else user.username

#     duplicate_report = None

#     # 2. Handle Form Submission (POST request)
#     if request.method == 'POST':
#         form = DailyReportForm(request.POST)
#         if form.is_valid():
#             year = form.cleaned_data.get('year')
#             volume_num = form.cleaned_data.get('volume_num')
#             location = form.cleaned_data.get('location')

#             # Check if an identical report already exists
#             existing = DailyReport.objects.filter(
#                 year=year, volume_num=volume_num, location=location
#             ).first()

#             if existing:
#                 duplicate_report = {
#                     'pk': existing.pk,
#                     'year': existing.year,
#                     'volume_num': existing.volume_num,
#                     'location': existing.get_location_display(),
#                     'name': existing.name,
#                     'created_at': existing.created_at,
#                 }

#                 return render(request, 'core/add_report.html', {
#                     'form': form, # Keeps their input data on screen
#                     'next': request.GET.get('next', ''),
#                     'auto_name': auto_name,
#                     'duplicate_report': duplicate_report,
#                 })

#             # Save the new report with the auto_name enforced
#             instance = form.save(commit=False)
#             instance.name = auto_name
#             instance.save()

#             # create the submodel
#             sync_workflow_records(instance, request.user)

#             # Success notification
#             messages.success(
#                 request, 
#                 f"Daily report for {instance.name} at {instance.get_location_display()} recorded successfully!"
#             )

#             # Determine where to redirect next
#             next_url = request.GET.get('next') or request.POST.get('next')
#             if next_url:
#                 return redirect(next_url)
#             return redirect(reverse('add_daily_report'))

#     # 3. Handle Initial Page Load (GET request)
#     else:
#         # Build initial data dictionary matching your CBV's get_initial logic
#         initial_data = {
#             'name': auto_name,
#         }
#         # for field in ['year', 'volume_num', 'location']:
#         #     val = request.GET.get(field)
#         #     if val:
#         #         initial_data[field] = val

#         form = DailyReportForm(initial=initial_data)

#     context = {
#         'form': form,
#         'next': request.GET.get('next', ''),
#         'auto_name': auto_name,
#     }
#     return render(request, 'core/add_report.html', context)


# ── DAILY REPORT: BULK ADD (login required) ──
@login_required
def add_multiple_reports_view(request):
    DailyReportFormSet = modelformset_factory(
        DailyReport,
        form=DailyReportBulkForm,
        extra=1,
        can_delete=False
    )

    # Auto-get name from logged-in user
    user = request.user
    full_name = user.get_full_name().strip()
    auto_name = full_name if full_name else user.username

    duplicate_reports = []

    if request.method == 'POST':
        # just like creating a form = xyzform(request.post)
        formset = DailyReportFormSet(request.POST)

        if formset.is_valid():
            instances = formset.save(commit=False)
            saved_count = 0
            duplicate_reports = []

            for instance in instances:
                # Always use logged-in user's name
                instance.name = auto_name
                existing = DailyReport.objects.filter(
                    year=instance.year,
                    volume_num=instance.volume_num,
                    location=instance.location
                ).first()

                if existing:
                    duplicate_reports.append({
                        'pk': existing.pk,
                        'year': existing.year,
                        'volume_num': existing.volume_num,
                        'location': existing.get_location_display(),
                        'name': existing.name,
                        'date': existing.date,
                    })
                else:
                    instance.save()
                    # create the submodels
                    sync_workflow_records(instance, request.user)
                    saved_count += 1

            if saved_count:
                messages.success(request, f"Successfully saved {saved_count} daily report(s) for {auto_name}.")

            if duplicate_reports:
                formset = DailyReportFormSet(queryset=DailyReport.objects.none())
                return render(request, 'core/add_multiple_reports.html', {
                    'formset': formset,
                    'duplicate_reports': duplicate_reports,
                    'auto_name': auto_name,
                })

            if saved_count:
                return redirect('add_daily_report_bulk')

    else:
        formset = DailyReportFormSet(queryset=DailyReport.objects.none())

    return render(request, 'core/add_multiple_reports.html', {
        'formset': formset,
        'duplicate_reports': duplicate_reports,
        'auto_name': auto_name,
    })


@login_required
def report_list_view(request):
    # Setup the base user role check right away
    is_admin = request.user.is_staff or request.user.is_superuser

    # Standardize user full name for filtering personal metrics
    user_full_name = request.user.get_full_name().upper() if request.user.get_full_name() else request.user.username.upper()

    # 🌟 LEVEL ACCESS GATEWAY 🌟 
    if is_admin:
        # Admins pull comprehensive datasets across the entire workforce 
        pdf_data = PDFRecord.objects.all()
        ind_data = IndexingRecord.objects.all()
        upload_data = UploadingRecord.objects.all()
        qc_data = QCRecord.objects.all()
        metadata_data = MetadataRecord.objects.all()
        
        # Admins see all logs out-of-the-box
        scanning_report = DailyReport.objects.all()
    else:
        # Regular employees are strictly limited to their own record contributions
        pdf_data = PDFRecord.objects.filter(created_by=request.user)
        ind_data = IndexingRecord.objects.filter(created_by=request.user)
        upload_data = UploadingRecord.objects.filter(created_by=request.user)
        qc_data = QCRecord.objects.filter(created_by=request.user)
        metadata_data = MetadataRecord.objects.filter(created_by=request.user)
        
        # Scanners only see their own names mapped on DailyReports
        scanning_report = DailyReport.objects.none()
        if request.user.username.endswith('-S'):
            scanning_report = DailyReport.objects.filter(name=user_full_name)

    # Collect unique DailyReport IDs from all active records 
    daily_report_ids = set(chain(
        pdf_data.values_list('daily_report_id', flat=True),
        ind_data.values_list('daily_report_id', flat=True),
        upload_data.values_list('daily_report_id', flat=True),
        qc_data.values_list('daily_report_id', flat=True),
        metadata_data.values_list('daily_report_id', flat=True),
    ))

    # Fetch corresponding DailyReport objects matching the sub-model IDs
    daily_reports = DailyReport.objects.filter(id__in=daily_report_ids)

    # NOTE : all the above data can be simply be created by 1 single query
    # daily_reports = DailyReport.objects.filter(
    #     Q(pdf_records__created_by=request.user) |
    #     Q(indexing_records__created_by=request.user) |
    #     Q(uploading_records__created_by=request.user) |
    #     Q(qc_records__created_by=request.user) |
    #     Q(metadata_records__created_by=request.user)
    # ).distinct()

    # Combine the datasets into a single unified workspace query
    final_report = (daily_reports | scanning_report).distinct()
    # print(final_report)
    # print(daily_reports)
    # print(final_report)

    # Build the admin employee dropdown before applying filters so the first
    # employee can be selected on the initial page load.
    employee_names = []
    if is_admin:
        report_names = DailyReport.objects.exclude(name='').values_list(
            'name', flat=True
        ).distinct()
        account_names = []
        for employee in User.objects.filter(
            is_active=True, is_staff=False, is_superuser=False
        ):
            full_name = employee.get_full_name().strip()
            account_names.append(
                (full_name if full_name else employee.username).upper()
            )
        employee_names = sorted(set(chain(report_names, account_names)))

    # Capture URL GET Filter Parameters
    filter_date = request.GET.get('date', '').strip()
    filter_month = request.GET.get('month', '').strip()
    filter_location = request.GET.get('location', '').strip()
    filter_stage = request.GET.get('workflow_stage', '').strip()

    # An explicit empty value means "All Employees". On the initial load, the
    # first real employee (the second dropdown option) is selected by default.
    selected_names = []
    if is_admin:
        if 'name' in request.GET:
            selected_names = [
                name.strip() for name in request.GET.getlist('name')
                if name.strip()
            ]
        elif employee_names:
            selected_names = [employee_names[0]]

    # Apply Standard Date Filter
    if filter_date:
        pdf_data = pdf_data.filter(created_at__date=filter_date)
        ind_data = ind_data.filter(created_at__date=filter_date)
        upload_data = upload_data.filter(created_at__date=filter_date)
        qc_data = qc_data.filter(created_at__date=filter_date) 
        metadata_data = metadata_data.filter(created_at__date=filter_date)

        # For admins, this filters all global daily reports by that date. For regular scanners, it filters just theirs.
        scanning_report = scanning_report.filter(date=filter_date) 
        

        daily_report_ids = set(chain(
            pdf_data.values_list('daily_report_id', flat=True),
            ind_data.values_list('daily_report_id', flat=True),
            upload_data.values_list('daily_report_id', flat=True),
            qc_data.values_list('daily_report_id', flat=True),
            metadata_data.values_list('daily_report_id', flat=True),
        ))

        daily_reports = DailyReport.objects.filter(id__in=daily_report_ids)
        final_report = (daily_reports | scanning_report).distinct()

    # Apply Standard Month Filter
    if filter_month:
        try:
            year_part, month_part = map(int, filter_month.split('-'))
            
            pdf_data = pdf_data.filter(created_at__year=year_part, created_at__month=month_part)
            ind_data = ind_data.filter(created_at__year=year_part, created_at__month=month_part)
            upload_data = upload_data.filter(created_at__year=year_part, created_at__month=month_part)
            qc_data = qc_data.filter(created_at__year=year_part, created_at__month=month_part)
            metadata_data = metadata_data.filter(created_at__year=year_part, created_at__month=month_part)

            scanning_report = scanning_report.filter(date__year=year_part, date__month=month_part)

            daily_report_ids = set(chain(
                pdf_data.values_list('daily_report_id', flat=True),
                ind_data.values_list('daily_report_id', flat=True),
                upload_data.values_list('daily_report_id', flat=True),
                qc_data.values_list('daily_report_id', flat=True),
                metadata_data.values_list('daily_report_id', flat=True),
            ))

            daily_reports = DailyReport.objects.filter(id__in=daily_report_ids)
            final_report = (daily_reports | scanning_report).distinct()

        except ValueError:
            pass

    if filter_location:
        final_report = final_report.filter(location=filter_location)

    # Match any selected employee across scanning and all workflow records.
    if selected_names and is_admin:
        if filter_stage == 'scanning':
            name_q = Q(name__in=selected_names)
        else:
            selected_user_ids = []
            for employee in User.objects.filter(is_active=True):
                full_name = employee.get_full_name().strip()
                display_name = (
                    full_name if full_name else employee.username
                ).upper()
                if display_name in selected_names:
                    selected_user_ids.append(employee.pk)

            name_q = (
                Q(name__in=selected_names) |
                Q(pdf_records__created_by_id__in=selected_user_ids) |
                Q(indexing_records__created_by_id__in=selected_user_ids) |
                Q(uploading_records__created_by_id__in=selected_user_ids) |
                Q(qc_records__created_by_id__in=selected_user_ids) |
                Q(metadata_records__created_by_id__in=selected_user_ids)
            )
        final_report = final_report.filter(name_q).distinct()

    # Workflow Sub-Model Stage Filtering
    # Refactored Workflow Sub-Model Stage Filtering

    if filter_stage:
        if filter_stage == 'scanning':
            if is_admin:
                # 1. Look up all User accounts whose username ends with '-S'
                scanner_usernames = User.objects.filter(username__endswith='-S')
                
                # 2. Extract their full names (or usernames) to match against DailyReport.name
                # We handle both upper/lower casting just in case
                scanner_names = []
                for u in scanner_usernames:
                    full_name = u.get_full_name().upper() if u.get_full_name() else u.username.upper()
                    scanner_names.append(full_name)
                    # Also include the lowercase/original version as a fallback
                    scanner_names.append(u.username)

                # 3. Filter final_report to only include items scanned by these accounts
                final_report = final_report.filter(name__in=scanner_names) 
            else:
                if request.user.username.endswith('-S'):
                    final_report = final_report.filter(name=user_full_name)
                else:
                    final_report = final_report.none()

            # add month and date filter on scanning report
            if filter_month:
                year_part, month_part = map(int, filter_month.split('-'))
                final_report = final_report.filter(created_at__year=year_part, created_at__month=month_part) 
            if filter_date : 
                final_report = final_report.filter(created_at__date=filter_date)

        elif filter_stage == 'pdf':
            if is_admin:
                # Admins see any daily report that has a PDF entry
                final_report = final_report.filter(pdf_records__isnull=False)
            else:
                # Regular users only see reports where THEY created the PDF entry
                final_report = final_report.filter(pdf_records__created_by=request.user)
                
        elif filter_stage == 'indexing':
            if is_admin:
                final_report = final_report.filter(indexing_records__isnull=False)
            else:
                final_report = final_report.filter(indexing_records__created_by=request.user)
                
        elif filter_stage == 'uploading':
            if is_admin:
                final_report = final_report.filter(uploading_records__isnull=False)
            else:
                final_report = final_report.filter(uploading_records__created_by=request.user)
                
        elif filter_stage == 'qc':
            if is_admin:
                final_report = final_report.filter(qc_records__isnull=False)
            else:
                final_report = final_report.filter(qc_records__created_by=request.user)
                
        elif filter_stage == 'metadata':
            if is_admin:
                final_report = final_report.filter(metadata_records__isnull=False)
            else:
                final_report = final_report.filter(metadata_records__created_by=request.user)

    final_report = final_report.order_by('-date', 'name')

    # Calculate Global Totals for the current workspace (Consolidated Metrics)
    totals = final_report.aggregate(
        sum_deeds=Sum('num_of_deed'),
        sum_pages=Sum('num_of_page'),
        sum_pdf=Sum('num_of_deed', filter=Q(pdf_deed=True)),
        sum_indexing=Sum('num_of_deed', filter=Q(indexing=True)),
        sum_uploading=Sum('num_of_deed', filter=Q(uploading=True)),
        sum_qc=Sum('num_of_deed', filter=Q(QC=True)),
        sum_metadata=Sum('num_of_page', filter=Q(metadata=True))
    )

    # Calculate User Metrics (Personal Performance Summary)
    # Calculate User Metrics (Personal Performance Summary)
    user_metrics = {
        # 1. Pages Scanned (Only tracks if user is a scanner account ending with -S)
        'pages_scanned_count': (
            final_report.filter(name__iexact=user_full_name).aggregate(s=Sum('num_of_page'))['s'] or 0
        ) if request.user.username.endswith('-S') else 0,
        
        # 2. PDF Created (Sum of deeds from daily reports where this user created the PDF work entry)
        'pdf_count': PDFRecord.objects.filter(
            created_by=request.user, 
            daily_report__in=final_report,
            daily_report__pdf_deed=True  # Confirms the PDF checkbox status is active
        ).aggregate(s=Sum('daily_report__num_of_deed'))['s'] or 0,
        
        # 3. Indexing Done (Sum of deeds from daily reports indexed by this user)
        'indexing_count': IndexingRecord.objects.filter(
            created_by=request.user, 
            daily_report__in=final_report,
            daily_report__indexing=True  # Confirms the Indexing checkbox status is active
        ).aggregate(s=Sum('daily_report__num_of_deed'))['s'] or 0,
        
        # 4. Uploading Done (Sum of deeds from daily reports uploaded by this user)
        'uploading_count': UploadingRecord.objects.filter(
            created_by=request.user, 
            daily_report__in=final_report,
            daily_report__uploading=True  # Confirms the Uploading checkbox status is active
        ).aggregate(s=Sum('daily_report__num_of_deed'))['s'] or 0,
        
        # 5. QC Verified (Sum of deeds from daily reports QC verified by this user)
        'qc_count': QCRecord.objects.filter(
            created_by=request.user, 
            daily_report__in=final_report,
            daily_report__QC=True          # Confirms the QC checkbox status is active
        ).aggregate(s=Sum('daily_report__num_of_deed'))['s'] or 0,
        
        # 6. Metadata Saved (Sum of pages from daily reports metadata processed by this user)
        'metadata_count': MetadataRecord.objects.filter(
            created_by=request.user, 
            daily_report__in=final_report,
            daily_report__metadata=True    # Confirms the Metadata checkbox status is active
        ).aggregate(s=Sum('daily_report__num_of_page'))['s'] or 0,
    }

    # Excel Export Engine
    if 'export_excel' in request.GET:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Location grouping for the two report tabs
        GAYA_LOCATIONS = ['SHERGHATI', 'TEKARI', 'AURANGABAD', 'JAHANABAD', 'NAWADA']
        gaya_reports = final_report.filter(location__in=GAYA_LOCATIONS)
        nalanda_reports = final_report.exclude(location__in=GAYA_LOCATIONS)

        wb = Workbook()
        wb.remove(wb.active)  # remove default blank sheet, we'll add our own named ones

        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F2935', end_color='1F2935', fill_type='solid')
        data_font = Font(name='Calibri', size=11, bold=False)

        summary_label_font = Font(name='Calibri', size=11, bold=True, color='1F2935')
        summary_val_font = Font(name='Calibri', size=11, bold=True, color='000000')
        summary_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')

        thin_side = Side(border_style="thin", color="D3D3D3")
        thick_bottom_side = Side(border_style="medium", color="1F2935")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        headers = ['Date', 'Location', 'Employee Name', 'Year', 'Volume No.',
                   'No. Deeds', 'No. Pages', 'PDF', 'Indexing', 'Uploading', 'QC', 'Metadata']

        def write_report_sheet(sheet_title, reports_qs):
            """Writes one location-grouped tab: header row, data rows, and its own summary block."""
            ws = wb.create_sheet(title=sheet_title)

            ws.append(headers)
            ws.row_dimensions[1].height = 24
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            for report in reports_qs:
                row_data = [
                    report.date.strftime('%d/%m/%Y') if report.date else '',
                    report.get_location_display(),
                    report.name,
                    report.year,
                    report.volume_num,
                    report.num_of_deed,
                    report.num_of_page,
                    "Yes" if report.pdf_deed else "No",
                    "Yes" if report.indexing else "No",
                    "Yes" if report.uploading else "No",
                    "Yes" if report.QC else "No",
                    "Yes" if report.metadata else "No",
                ]
                ws.append(row_data)

                current_row = ws.max_row
                ws.row_dimensions[current_row].height = 20
                for col_idx, cell in enumerate(ws[current_row], start=1):
                    cell.font = data_font
                    cell.border = thin_border
                    if col_idx in [1, 4, 5, 6, 7, 8, 9, 10, 11, 12]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align

            ws.append([])  # Spacer

            # Per-tab totals (scoped to this location group only)
            tab_totals = reports_qs.aggregate(
                sum_deeds=Sum('num_of_deed'),
                sum_pages=Sum('num_of_page'),
                sum_pdf=Sum('num_of_deed', filter=Q(pdf_deed=True)),
                sum_indexing=Sum('num_of_deed', filter=Q(indexing=True)),
                sum_uploading=Sum('num_of_deed', filter=Q(uploading=True)),
                sum_qc=Sum('num_of_deed', filter=Q(QC=True)),
                sum_metadata=Sum('num_of_page', filter=Q(metadata=True))
            )

            if is_admin:
                summary_rows = [
                    ("Total Number of volume", reports_qs.count()),
                    ("Total Deeds", tab_totals['sum_deeds'] or 0),
                    ("Total Pages", tab_totals['sum_pages'] or 0),
                    ("Total PDF Created", tab_totals['sum_pdf'] or 0),
                    ("Total Indexing", tab_totals['sum_indexing'] or 0),
                    ("Total Uploading", tab_totals['sum_uploading'] or 0),
                    ("Total QC", tab_totals['sum_qc'] or 0),
                    ("Total Metadata", tab_totals['sum_metadata'] or 0),
                ]
            else:
                summary_rows = [
                    ("Total Number of volume", reports_qs.count()),
                    ("Total Pages", tab_totals['sum_pages'] or 0),
                    ("Total Deed", tab_totals['sum_deeds'] or 0),
                ]

            for label, val in summary_rows:
                ws.append(["", "", "", "", label, val])
                s_row = ws.max_row
                ws.row_dimensions[s_row].height = 22

                lbl_cell = ws.cell(row=s_row, column=5)
                lbl_cell.font = summary_label_font
                lbl_cell.alignment = right_align
                lbl_cell.fill = summary_fill
                lbl_cell.border = Border(left=thin_side, top=thin_side, bottom=thin_side)

                val_cell = ws.cell(row=s_row, column=6)
                val_cell.font = summary_val_font
                val_cell.alignment = center_align
                val_cell.fill = summary_fill
                val_cell.border = Border(right=thin_side, top=thin_side, bottom=thin_side)

                if label == summary_rows[-1][0]:
                    lbl_cell.border = Border(left=thin_side, top=thin_side, bottom=thick_bottom_side)
                    val_cell.border = Border(right=thin_side, top=thin_side, bottom=thick_bottom_side)

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Build the two tabs
        write_report_sheet("Gaya Report", gaya_reports)
        write_report_sheet("Nalanda Report", nalanda_reports)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Reports_Export_{datetime.date.today()}.xlsx'
        wb.save(response)
        return response


    # Pagination (totals above are computed on the full filtered queryset)
    total_volume = final_report.count()
    paginator = Paginator(final_report, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_range = paginator.get_elided_page_range(page_obj.number, on_each_side=1, on_ends=1)

    # Preserve active filters in pagination links
    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    # Context Preparation
    existing_dates = DailyReport.objects.dates('date', 'month', order='DESC')
    available_months = [
        {'value': d.strftime('%Y-%m'), 'display': d.strftime('%B %Y')}
        for d in existing_dates
    ]

    workflow_stages = [
        {'value': 'scanning', 'display': 'Scanning Report'},
        {'value': 'pdf', 'display': 'PDF Report'},
        {'value': 'indexing', 'display': 'Indexing Report'},
        {'value': 'uploading', 'display': 'Uploading Report'},
        {'value': 'qc', 'display': 'QC Report'},
        {'value': 'metadata', 'display': 'Metadata Report'},
    ]

    ends_with_s = str(request.user.username).endswith('-S')

    context = {
        'reports': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'page_range': page_range,
        'querystring': querystring,
        'total_volume': total_volume,
        'employee_names': employee_names,
        'total_deeds': totals['sum_deeds'],
        'total_pages': totals['sum_pages'],
        'total_pdf': totals['sum_pdf'],          
        'total_indexing': totals['sum_indexing'], 
        'total_uploading': totals['sum_uploading'],
        'total_qc': totals['sum_qc'],            
        'total_metadata': totals['sum_metadata'],    
        'user_metrics': user_metrics,
        'locations': DailyReport.LOCATION_CHOICES,
        'available_months': available_months,
        'workflow_stages': workflow_stages,       
        'selected_date': filter_date,
        'selected_month': filter_month,
        'selected_location': filter_location,
        'selected_names': selected_names,
        'selected_name': selected_names[0] if selected_names else '',
        'selected_employee_label': (
            ', '.join(selected_names) if selected_names else 'All Employees'
        ),
        'selected_stage': filter_stage,           
        'is_admin': is_admin,
        'ends_with_s' : ends_with_s,
    }
    return render(request, 'core/report_list.html', context)


# ── UPDATE REPORT SEARCH (login required) ──
@login_required
def update_report_search_view(request):
    report = None
    not_found = False
    searched = False

    year = request.GET.get('year', '').strip()
    volume_num = request.GET.get('volume_num', '').strip()
    location = request.GET.get('location', '').strip()

    if year and volume_num and location:
        searched = True
        try:
            report = DailyReport.objects.get(
                year=year, volume_num=volume_num, location=location
            )
        except DailyReport.DoesNotExist:
            not_found = True
        except DailyReport.MultipleObjectsReturned:
            report = DailyReport.objects.filter(
                year=year, volume_num=volume_num, location=location
            ).order_by('-date').first()

    context = {
        'report': report,
        'not_found': not_found,
        'searched': searched,
        'year': year,
        'volume_num': volume_num,
        'location': location,
        'locations': DailyReport.LOCATION_CHOICES,
    }
    return render(request, 'core/update_report_search.html', context)

@login_required
def daily_report_update_view(request, pk):
    # 1. Fetch the existing report instance 
    report = get_object_or_404(DailyReport, pk=pk)
    user = request.user

    # data for submodels 
    full_name = user.get_full_name().strip()
    current_updater_name = full_name if full_name else user.username
    
    # Define our fields 
    BOOLEAN_FIELDS = ['num_of_deed','scanning', 'pdf_deed', 'indexing', 'uploading', 'QC', 'metadata']
    PROTECTED_FIELDS = ['date', 'location', 'name', 'year', 'volume_num', 'num_of_page']
    
    # Check if the record is older than 24 hours
    is_expired = timezone.now() > report.created_at + timedelta(hours=24)

    # 1. Initialize the form based on request method
    if request.method == 'POST':
        form = DailyReportUpdateForm(request.POST, instance=report)
    else:
        form = DailyReportUpdateForm(instance=report)

    # 2. 🌟 APPLY DISABLING LOGIC HERE (Runs for BOTH GET and POST) 🌟
    if is_expired and not user.is_superuser:
        for field_name, field in form.fields.items():
            if field_name not in BOOLEAN_FIELDS:
                field.disabled = True

    # 2. Handle Form Submission (POST)
    if request.method == 'POST':
        
        if form.is_valid():
            # Enforce 24-hour rule protection on post-back for non-superusers
            if is_expired and not user.is_superuser:
                original = DailyReport.objects.get(pk=report.pk)
                for field in PROTECTED_FIELDS:
                    setattr(form.instance, field, getattr(original, field)) # checkout this one
            
            updated_report = form.save()

            # create the submodels
            sync_workflow_records(updated_report, request.user)


            messages.success(request, "Report updated successfully!")
            
            # Read the hidden tracking URL sent by the form
            back_to_url = request.POST.get('back_to_url')
            if back_to_url:
                return redirect(back_to_url)
                
            # Default fallback url matching your old get_success_url
            return redirect(
                reverse('update_report_search') + 
                f"?year={report.year}&volume_num={report.volume_num}&location={report.location}"
            )
        
        # Apply field disabling logic if older than 24 hours (and not an admin)
        if is_expired and not user.is_superuser:
            for field_name, field in form.fields.items():
                if field_name not in BOOLEAN_FIELDS:
                    field.disabled = True

    # 4. Handle referer url tracking context
    if request.method == 'GET':
        back_to_url = request.META.get('HTTP_REFERER', '')
    else:
        back_to_url = request.POST.get('back_to_url', '')

    context = {
        'form': form,
        'report': report,
        'back_to_url': back_to_url,
    }
    
    return render(request, 'core/update_report.html', context)