import csv
from datetime import timedelta
from io import BytesIO
from urllib.parse import urlencode

from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin, GroupAdmin as BaseGroupAdmin
from django.contrib.auth.models import User, Group
from django.db import models
from django.http import HttpResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.safestring import mark_safe

from unfold.admin import ModelAdmin, TabularInline
from unfold.forms import AdminPasswordChangeForm, UserChangeForm, UserCreationForm
from unfold.contrib.filters.admin import (
    ChoicesDropdownFilter,
    DropdownFilter,
    RangeDateFilter,
    RelatedDropdownFilter,
)
from .models import ContactMessage, CompanyCertificate, DailyReport, PDFRecord, IndexingRecord, UploadingRecord, QCRecord, MetadataRecord
from .utils import sync_workflow_records


# ── Daily Report export helpers ────────────────────────────
DAILY_REPORT_EXPORT_HEADERS = [
    "Date",
    "Employee",
    "District",
    "Location",
    "Year",
    "Volume",
    "Pages",
    "Deeds",
    "Scanning",
    "PDF",
    "Indexing",
    "Uploading",
    "QC",
    "Metadata",
    "Created at",
]

# Material-style document SVG for PDF stage (fill via CSS currentColor)
PDF_STAGE_SVG = (
    '<svg class="ngs-dr-stage-svg" xmlns="http://www.w3.org/2000/svg" '
    'viewBox="0 -960 960 960" width="16" height="16" fill="currentColor" aria-hidden="true">'
    '<path d="M280-80q-33 0-56.5-23.5T200-160v-640q0-33 23.5-56.5T280-880h247q16 0 '
    '30.5 6t25.5 17l154 154q11 11 17 25.5t6 30.5v487q0 33-23.5 56.5T680-80H280Zm160-720H280v640'
    'h400v-400H560q-50 0-85-35t-35-85v-120Zm80 0v120q0 17 11.5 28.5T560-640h120v-7L527-800h-7Z'
    'M400-200q-17 0-28.5-11.5T360-240q0-17 11.5-28.5T400-280h80q17 0 28.5 11.5T520-240q0 17'
    '-11.5 28.5T480-200h-80Zm0-160q-17 0-28.5-11.5T360-400q0-17 11.5-28.5T400-440h160q17 0 '
    '28.5 11.5T600-400q0 17-11.5 28.5T560-360H400Z"/>'
    '</svg>'
)


def _yn(value):
    return "Yes" if bool(value) else "No"


def _dr_row(report):
    return [
        report.date.strftime("%Y-%m-%d") if report.date else "",
        report.name or "",
        report.get_district_display() if report.district else "",
        report.get_location_display() if report.location else "",
        report.year or "",
        report.volume_num or "",
        report.num_of_page if report.num_of_page is not None else "",
        report.num_of_deed if report.num_of_deed is not None else "",
        _yn(report.scanning),
        _yn(report.pdf_deed),
        _yn(report.indexing),
        _yn(report.uploading),
        _yn(report.QC),
        _yn(report.metadata),
        timezone.localtime(report.created_at).strftime("%Y-%m-%d %H:%M") if report.created_at else "",
    ]


def _dr_iter_rows(queryset):
    for report in queryset.iterator(chunk_size=500):
        yield _dr_row(report)


def _dr_filename(ext):
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    return f"daily-reports-{stamp}.{ext}"


def export_daily_reports_csv(queryset):
    """Stream a filtered DailyReport queryset as CSV."""
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{_dr_filename("csv")}"'
    # Excel-friendly UTF-8 BOM so Unicode names render correctly
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(DAILY_REPORT_EXPORT_HEADERS)
    for row in _dr_iter_rows(queryset):
        writer.writerow(row)
    return response


def export_daily_reports_xlsx(queryset):
    """Return the filtered DailyReport queryset as an .xlsx download."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # write_only keeps memory low on production (2k+ rows)
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Daily Reports")

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF16A34A")
    center = Alignment(horizontal="center", vertical="center")

    header_cells = []
    from openpyxl.cell import WriteOnlyCell
    for title in DAILY_REPORT_EXPORT_HEADERS:
        cell = WriteOnlyCell(sheet, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        header_cells.append(cell)
    sheet.append(header_cells)

    for row in _dr_iter_rows(queryset):
        sheet.append(row)

    # Remove the default empty sheet openpyxl creates
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    buf = BytesIO()
    workbook.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_dr_filename("xlsx")}"'
    return response

admin.site.index_title = "Dashboard"


class PremiumAdmin(ModelAdmin):
    """Shared responsive list controls for all admin pages."""

    list_per_page = 25
    list_max_show_all = 100
    list_filter_submit = True
    list_filter_sheet = False
    show_full_result_count = False


# ── USERS & GROUPS (restyled with the Unfold theme) ──
admin.site.unregister(User)
admin.site.unregister(Group)


@admin.register(User)
class UserAdmin(BaseUserAdmin, ModelAdmin):
    # Unfold-styled forms so the add/change user pages match the admin theme
    form = UserChangeForm
    add_form = UserCreationForm
    change_password_form = AdminPasswordChangeForm

    list_display = ('username', 'full_name_display', 'role_display', 'is_active', 'last_login')
    list_filter = ('is_staff', 'is_superuser', 'is_active')
    search_fields = ('username', 'first_name', 'last_name', 'email')
    search_help_text = "Search by username, name, or email"
    ordering = ('username',)
    list_per_page = 25
    list_filter_submit = True
    list_filter_sheet = False

    add_fieldsets = (
        ('Account Credentials', {
            'classes': ('wide',),
            'fields': ('username', 'password1', 'password2'),
            'description': "Usernames ending with -S are treated as scanner accounts.",
        }),
        ('Employee Details', {
            'classes': ('wide',),
            'fields': ('first_name', 'last_name', 'email'),
            'description': "First and last name appear as the employee name on daily reports.",
        }),
        ('Permissions', {
            'classes': ('wide',),
            'fields': ('is_active', 'is_staff', 'is_superuser'),
            'description': "Staff status makes this user an admin who can see all reports.",
        }),
    )

    fieldsets = (
        ('Account Credentials', {'fields': ('username', 'password')}),
        ('Employee Details', {'fields': ('first_name', 'last_name', 'email')}),
        ('Permissions', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        ('Important Dates', {'fields': ('last_login', 'date_joined'), 'classes': ('collapse',)}),
    )

    def full_name_display(self, obj):
        return obj.get_full_name() or "—"
    full_name_display.short_description = "Full Name"

    def role_display(self, obj):
        if obj.is_superuser:
            return format_html('<strong style="color:#dc2626;">Superuser</strong>')
        if obj.is_staff:
            return format_html('<strong style="color:#d97706;">Admin</strong>')
        if obj.username.endswith('-S'):
            return format_html('<strong style="color:#15803d;">Scanner</strong>')
        return "Employee"
    role_display.short_description = "Role"


@admin.register(Group)
class GroupAdmin(BaseGroupAdmin, ModelAdmin):
    pass


class ReportYearDropdownFilter(DropdownFilter):
    title = "year"
    parameter_name = "year"

    def lookups(self, request, model_admin):
        years = DailyReport.objects.order_by("-year").values_list("year", flat=True).distinct()
        return [(year, year) for year in years if year]

    def queryset(self, request, queryset):
        return queryset.filter(year=self.value()) if self.value() else queryset


class RelatedReportYearDropdownFilter(ReportYearDropdownFilter):
    parameter_name = "daily_report__year"

    def queryset(self, request, queryset):
        return (
            queryset.filter(daily_report__year=self.value())
            if self.value()
            else queryset
        )


class CertificateTypeDropdownFilter(DropdownFilter):
    title = "certificate title"
    parameter_name = "certificate_name"

    def lookups(self, request, model_admin):
        names = (
            CompanyCertificate.objects.order_by("certificate_name")
            .values_list("certificate_name", flat=True)
            .distinct()
        )
        return [(name, name) for name in names if name]

    def queryset(self, request, queryset):
        return (
            queryset.filter(certificate_name=self.value())
            if self.value()
            else queryset
        )


@admin.register(ContactMessage)
class ContactMessageAdmin(PremiumAdmin):
    list_display = ('name', 'email', 'phone', 'subject', 'created_at', 'is_read')
    list_filter = ('is_read', ('created_at', RangeDateFilter))
    search_fields = ('name', 'email', 'subject')
    search_help_text = "Search by name, email, or subject"
    readonly_fields = ('created_at',)
    actions = ['mark_as_read']

    def mark_as_read(self, request, queryset):
        queryset.update(is_read=True)
    mark_as_read.short_description = "Mark selected as read"



@admin.register(CompanyCertificate)
class CompanyCertificateAdmin(PremiumAdmin):
    # Removed 'expiry_date' from the end of this list
    list_display = ('certificate_number', 'recipient_name', 'certificate_name', 'issue_date')
    
    list_display_links = ('certificate_number', 'recipient_name')
    search_fields = ('recipient_name', 'certificate_number', 'certificate_name')
    search_help_text = "Search by recipient, certificate number, or title"
    list_filter = (
        CertificateTypeDropdownFilter,
        ('issue_date', RangeDateFilter),
    )
    
    # Cleaned up the fieldsets layout to remove the dates breakdown
    fieldsets = (
        ('Certificate Identification', {
            'fields': ('certificate_number', 'certificate_name')
        }),
        ('Recipient Information', {
            'fields': ('recipient_name',)
        }),
        ('Issue Date', {
            'fields': ('issue_date',)
        }),
    )

# --- Optional: Tabular Inlines ---
# This lets administrators view and edit the sub-records directly at the bottom
# of the DailyReport page without leaving the screen!
class PDFRecordInline(TabularInline):
    model = PDFRecord
    extra = 0
    fields = ('created_by', 'name', 'created_at')

class IndexingRecordInline(TabularInline):
    model = IndexingRecord
    extra = 0
    fields = ('created_by', 'name', 'created_at')

class UploadingRecordInline(TabularInline):
    model = UploadingRecord
    extra = 0
    fields = ('created_by', 'name', 'created_at')

class QCRecordInline(TabularInline):
    model = QCRecord
    extra = 0
    fields = ('created_by', 'name', 'created_at')

class MetadataRecordInline(TabularInline):
    model = MetadataRecord
    extra = 0
    fields = ('created_by', 'name', 'created_at')


# --- Main DailyReport Admin ---
@admin.register(DailyReport)
class DailyReportAdmin(PremiumAdmin):
    change_list_template = "admin/core/dailyreport_change_list.html"
    list_before_template = "admin/core/dailyreport_list_header.html"
    list_fullwidth = True
    list_display = (
        'serial_no',
        'date',
        'name',
        'district_badge',
        'archive_location',
        'deed_cell',
        'page_cell',
        'progress_cell',
        'pdf_badge',
        'indexing_badge',
        'uploading_badge',
        'qc_badge',
        'edit_button',
    )
    list_display_links = ('name',)
    list_filter = (
        ('date', RangeDateFilter),
        ('district', ChoicesDropdownFilter),
        ('location', ChoicesDropdownFilter),
        ReportYearDropdownFilter,
        'pdf_deed',
        'indexing',
        'uploading',
        'QC',
        'metadata',
    )
    # Open filters in an off-canvas drawer (opened via the hero button)
    list_filter_sheet = True
    list_filter_submit = True
    search_fields = ('name', 'year', 'volume_num', 'district', 'location')
    search_help_text = "Search by employee, year, volume, district, or location"

    readonly_fields = ('district',)

    actions = ("export_selected_as_csv", "export_selected_as_xlsx")

    # Embed the inline tables nicely at the bottom of the edit layout
    inlines = [PDFRecordInline, IndexingRecordInline, UploadingRecordInline, QCRecordInline, MetadataRecordInline]

    _STAGE_LABEL = {True: "Done", False: "—"}

    class Media:
        css = {"all": ("core/css/admin-premium.css",)}
        js = ("core/js/admin-filters-drawer.js",)

    # ── Custom columns ─────────────────────────────────────────
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # Stash the request so serial_no can compute against the current page
        self._current_request = request
        return qs

    def serial_no(self, obj):
        request = getattr(self, "_current_request", None)
        page = 1
        per_page = self.list_per_page
        if request is not None:
            try:
                page = max(int(request.GET.get("p", 0)) + 1, 1)
            except (TypeError, ValueError):
                page = 1
        counter = getattr(self, "_row_counter", 0) + 1
        self._row_counter = counter
        # Reset when a new page loads (row 1 for that page)
        base = (page - 1) * per_page
        return format_html(
            '<span class="ngs-dr-sl">{}</span>',
            base + counter if counter <= per_page else counter,
        )
    serial_no.short_description = "Sl no"

    def district_badge(self, obj):
        label = obj.get_district_display() if obj.district else "—"
        return format_html(
            '<span class="ngs-dr-district">{}</span>',
            label,
        )
    district_badge.short_description = "Dist"
    district_badge.admin_order_field = "district"

    def archive_location(self, obj):
        town = obj.get_location_display() if obj.location else "—"
        year = obj.year or "—"
        volume = obj.volume_num or "—"
        return format_html(
            '<div class="ngs-dr-loc">'
            '<span class="ngs-dr-loc-year">{}</span>'
            '<span class="ngs-dr-loc-vol">{} · Vol {}</span>'
            '</div>',
            town,
            year,
            volume,
        )
    archive_location.short_description = "Location"
    archive_location.admin_order_field = "location"

    def deed_cell(self, obj):
        val = obj.num_of_deed
        if val is None:
            return format_html('<span class="ngs-dr-num ngs-dr-num-empty">—</span>')
        return format_html('<span class="ngs-dr-num">{}</span>', f"{val:,}")
    deed_cell.short_description = "Deed"
    deed_cell.admin_order_field = "num_of_deed"

    def page_cell(self, obj):
        val = obj.num_of_page
        if val is None:
            return format_html('<span class="ngs-dr-num ngs-dr-num-empty">—</span>')
        return format_html('<span class="ngs-dr-num">{}</span>', f"{val:,}")
    page_cell.short_description = "Page"
    page_cell.admin_order_field = "num_of_page"

    def _stage_badge(self, done, icon, label, svg=None):
        cls = "ngs-dr-stage ngs-dr-stage-done" if done else "ngs-dr-stage ngs-dr-stage-todo"
        title = f"{label} — {'Done' if done else 'Pending'}"
        if svg:
            return format_html(
                '<span class="{}" title="{}">{}</span>',
                cls,
                title,
                mark_safe(svg),
            )
        return format_html(
            '<span class="{}" title="{}">'
            '<span class="material-symbols-outlined ngs-dr-stage-icon" aria-hidden="true">{}</span>'
            '</span>',
            cls,
            title,
            icon,
        )

    def pdf_badge(self, obj):
        return self._stage_badge(obj.pdf_deed, "docs", "PDF", svg=PDF_STAGE_SVG)
    pdf_badge.short_description = "PDF"
    pdf_badge.admin_order_field = "pdf_deed"

    def indexing_badge(self, obj):
        return self._stage_badge(obj.indexing, "list", "Indexing")
    indexing_badge.short_description = "Idx"
    indexing_badge.admin_order_field = "indexing"

    def uploading_badge(self, obj):
        return self._stage_badge(obj.uploading, "database_upload", "Uploading")
    uploading_badge.short_description = "Up"
    uploading_badge.admin_order_field = "uploading"

    def qc_badge(self, obj):
        return self._stage_badge(obj.QC, "list_alt_check", "QC")
    qc_badge.short_description = "QC"
    qc_badge.admin_order_field = "QC"

    def progress_cell(self, obj):
        stages = [obj.pdf_deed, obj.indexing, obj.uploading, obj.QC, obj.metadata]
        done = sum(1 for s in stages if s)
        total = len(stages)
        pct = int(round((done / total) * 100)) if total else 0
        if pct >= 100:
            tone = "ngs-dr-progress-done"
        elif pct >= 60:
            tone = "ngs-dr-progress-hi"
        elif pct >= 20:
            tone = "ngs-dr-progress-mid"
        else:
            tone = "ngs-dr-progress-lo"
        return format_html(
            '<div class="ngs-dr-progress {tone}" title="{done} of {total} stages complete">'
            '  <div class="ngs-dr-progress-bar"><span style="width:{pct}%;"></span></div>'
            '  <span class="ngs-dr-progress-pct">{pct}%</span>'
            '</div>',
            tone=tone, done=done, total=total, pct=pct,
        )
    progress_cell.short_description = "%"

    def edit_button(self, obj):
        """Compact icon-only edit control."""
        url = reverse('admin:core_dailyreport_change', args=[obj.pk])
        return format_html(
            '<a href="{}" class="ngs-dr-edit" title="Edit">'
            '<span class="material-symbols-outlined">edit</span></a>',
            url,
        )
    edit_button.short_description = ""

    # ── KPI + pipeline stats for the header ────────────────────
    def changelist_view(self, request, extra_context=None):
        self._row_counter = 0
        extra_context = extra_context or {}

        # Use the *filtered* queryset (respects active filters + search) so the
        # header KPIs, chips, and empty state all mirror what the operator sees.
        try:
            qs = self._filtered_queryset_from_request(request)
        except Exception:
            qs = self.get_queryset(request)
        total = qs.count()
        agg = qs.aggregate(
            pages=models.Sum("num_of_page"),
            deeds=models.Sum("num_of_deed"),
        )
        pages = agg["pages"] or 0
        deeds = agg["deeds"] or 0
        districts = qs.exclude(district="").values("district").distinct().count()
        pdf_done = qs.filter(pdf_deed=True).count()
        idx_done = qs.filter(indexing=True).count()
        up_done = qs.filter(uploading=True).count()
        qc_done = qs.filter(QC=True).count()
        meta_done = qs.filter(metadata=True).count()

        def _pct(part):
            return round((part / total) * 100) if total else 0

        district_label_map = dict(DailyReport.DISTRICT_CHOICES)
        district_rows = (
            qs.exclude(district="")
            .values("district")
            .annotate(count=models.Count("id"))
            .order_by("-count")
        )
        district_chips = [
            {
                "code": row["district"],
                "label": district_label_map.get(row["district"], row["district"].title()),
                "count": row["count"],
                "pct": _pct(row["count"]),
            }
            for row in district_rows
        ]

        extra_context["dr_kpis"] = [
            {"label": "Reports", "value": f"{total:,}", "hint": "volumes", "icon": "description"},
            {"label": "Pages", "value": f"{pages:,}", "hint": "scanned", "icon": "menu_book"},
            {"label": "Deeds", "value": f"{deeds:,}", "hint": "recorded", "icon": "folder_open"},
            {"label": "Districts", "value": f"{districts:,}", "hint": "active", "icon": "location_on"},
            {"label": "PDF done", "value": f"{_pct(pdf_done)}%", "hint": f"{pdf_done:,} of {total:,}", "icon": "picture_as_pdf"},
            {"label": "QC done", "value": f"{_pct(qc_done)}%", "hint": f"{qc_done:,} of {total:,}", "icon": "verified"},
        ]
        extra_context["dr_pipeline"] = [
            {"label": "PDF", "done": pdf_done, "pct": _pct(pdf_done), "icon": "picture_as_pdf"},
            {"label": "Indexing", "done": idx_done, "pct": _pct(idx_done), "icon": "toc"},
            {"label": "Uploading", "done": up_done, "pct": _pct(up_done), "icon": "cloud_upload"},
            {"label": "QC", "done": qc_done, "pct": _pct(qc_done), "icon": "task_alt"},
            {"label": "Metadata", "done": meta_done, "pct": _pct(meta_done), "icon": "bookmark_added"},
        ]
        extra_context["dr_district_chips"] = district_chips
        # Dropdown options for all three districts (counts from filtered set when possible)
        chip_map = {c["code"]: c for c in district_chips}
        # Prefer unfiltered totals for the district picker so users always see all options
        try:
            base_qs = self.get_queryset(request)
            all_counts = {
                row["district"]: row["count"]
                for row in base_qs.exclude(district="")
                .values("district")
                .annotate(count=models.Count("id"))
            }
        except Exception:
            all_counts = {c["code"]: c["count"] for c in district_chips}

        district_options = []
        for code, label in DailyReport.DISTRICT_CHOICES:
            district_options.append({
                "code": code,
                "label": label,
                "count": all_counts.get(code, 0),
                "active": request.GET.get("district__exact") == code,
                "url": self._param_url(request, patch={"district__exact": code}),
            })
        extra_context["dr_district_options"] = district_options
        extra_context["dr_district_all_url"] = self._param_url(request, drop=["district__exact"])
        extra_context["dr_add_url"] = reverse("admin:core_dailyreport_add")
        extra_context["dr_total"] = total

        # Count active filter params (excluding pagination / ordering / search)
        _skip = {"p", "o", "all", "e", "_changelist_filters"}
        extra_context["dr_active_filters"] = sum(
            1 for key, value in request.GET.items()
            if key not in _skip and value not in ("", None)
        )

        # Human-friendly chips for every active filter — one-click remove per chip
        extra_context["dr_active_filter_chips"] = self._build_active_filter_chips(request)

        # Quick date-range presets that keep other filters intact
        extra_context["dr_range_presets"] = self._build_range_presets(request)

        # Stage quick-toggles for the search toolbar (PDF / Indexing / Uploading / QC)
        extra_context["dr_stage_toggles"] = self._build_stage_toggles(request)

        # Clear-all URL preserves ordering only (drops filters + search + pagination)
        keep = {}
        if request.GET.get("o"):
            keep["o"] = request.GET["o"]
        extra_context["dr_clear_url"] = f"{request.path}?{urlencode(keep)}" if keep else request.path

        # One-click export URLs preserve the current filter + search context
        export_qs = request.GET.urlencode()
        try:
            base_csv = reverse("admin:core_dailyreport_export_csv")
            base_xlsx = reverse("admin:core_dailyreport_export_xlsx")
        except Exception:
            base_csv = "/admin/core/dailyreport/export/csv/"
            base_xlsx = "/admin/core/dailyreport/export/xlsx/"
        extra_context["dr_export_csv_url"] = f"{base_csv}?{export_qs}" if export_qs else base_csv
        extra_context["dr_export_xlsx_url"] = f"{base_xlsx}?{export_qs}" if export_qs else base_xlsx

        return super().changelist_view(request, extra_context)

    # ── Helpers powering the header UX ────────────────────────
    _BOOLEAN_LABELS = {"1": "Done", "0": "Pending", "True": "Done", "False": "Pending"}

    def _param_url(self, request, drop=None, patch=None):
        """Build a same-page URL with some params removed / added."""
        params = request.GET.copy()
        for key in (drop or ()):
            params.pop(key, None)
        for key, value in (patch or {}).items():
            if value is None:
                params.pop(key, None)
            else:
                params[key] = value
        # Reset pagination when the filter set changes
        params.pop("p", None)
        qs = params.urlencode()
        return f"{request.path}?{qs}" if qs else request.path

    def _build_active_filter_chips(self, request):
        district_labels = dict(DailyReport.DISTRICT_CHOICES)
        location_labels = dict(DailyReport.LOCATION_CHOICES)
        specs = [
            ("q", "Search", None, None),
            ("date_from", "From", None, None),
            ("date_to", "To", None, None),
            ("district__exact", "District", district_labels, None),
            ("location__exact", "Location", location_labels, None),
            ("year", "Year", None, None),
            ("pdf_deed__exact", "PDF", self._BOOLEAN_LABELS, None),
            ("indexing__exact", "Indexing", self._BOOLEAN_LABELS, None),
            ("uploading__exact", "Uploading", self._BOOLEAN_LABELS, None),
            ("QC__exact", "QC", self._BOOLEAN_LABELS, None),
            ("metadata__exact", "Metadata", self._BOOLEAN_LABELS, None),
        ]
        chips = []
        for key, title, value_map, _icon in specs:
            raw = request.GET.get(key)
            if raw in (None, "", "unknown"):
                continue
            label = value_map.get(raw, raw) if value_map else raw
            chips.append({
                "field": title,
                "value": label,
                "remove_url": self._param_url(request, drop=[key]),
            })
        return chips

    def _build_range_presets(self, request):
        today = timezone.localdate()
        current_from = request.GET.get("date_from")

        def preset(label, days=None, since=None):
            date_val = None
            if days is not None:
                date_val = (today - timedelta(days=days)).isoformat()
            elif since is not None:
                date_val = since.isoformat()
            active = (date_val or "") == (current_from or "")
            url = self._param_url(
                request,
                drop=["date_to"] if date_val is None else [],
                patch={"date_from": date_val},
            )
            return {"label": label, "url": url, "active": active}

        year_start = today.replace(month=1, day=1)
        return [
            preset("All", days=None),
            preset("Today", days=0),
            preset("7 days", days=6),
            preset("30 days", days=29),
            preset("This year", since=year_start),
        ]

    def _build_stage_toggles(self, request):
        """Icon toggles: click = filter Done; click again = clear that filter."""
        stages = [
            ("pdf_deed__exact", "PDF", "docs", PDF_STAGE_SVG),
            ("indexing__exact", "Indexing", "list", None),
            ("uploading__exact", "Uploading", "database_upload", None),
            ("QC__exact", "QC", "list_alt_check", None),
        ]
        toggles = []
        for key, label, icon, svg in stages:
            active = request.GET.get(key) == "1"
            if active:
                url = self._param_url(request, drop=[key])
            else:
                url = self._param_url(request, patch={key: "1"})
            toggles.append({
                "key": key,
                "label": label,
                "icon": icon,
                "svg": mark_safe(svg) if svg else None,
                "active": active,
                "url": url,
            })
        return toggles

    # ── Bulk export actions (selected rows) ────────────────────
    @admin.action(description="Export selected reports as CSV")
    def export_selected_as_csv(self, request, queryset):
        return export_daily_reports_csv(queryset)

    @admin.action(description="Export selected reports as Excel (.xlsx)")
    def export_selected_as_xlsx(self, request, queryset):
        return export_daily_reports_xlsx(queryset)

    # ── One-click filtered export (uses the current changelist filters) ──
    def get_urls(self):
        urls = super().get_urls()
        info = (self.model._meta.app_label, self.model._meta.model_name)
        custom = [
            path(
                "export/csv/",
                self.admin_site.admin_view(self.export_filtered_csv_view),
                name="%s_%s_export_csv" % info,
            ),
            path(
                "export/xlsx/",
                self.admin_site.admin_view(self.export_filtered_xlsx_view),
                name="%s_%s_export_xlsx" % info,
            ),
        ]
        return custom + urls

    def _filtered_queryset_from_request(self, request):
        """Reuse Django admin's ChangeList so exports honour every active filter/search."""
        try:
            # Preferred: Django's own ChangeList builder (safe across Unfold versions)
            cl = self.get_changelist_instance(request)
            return cl.get_queryset(request)
        except Exception:
            # Fallback: unfiltered base queryset — never 500 the download button
            return self.get_queryset(request)

    def export_filtered_csv_view(self, request):
        try:
            qs = self._filtered_queryset_from_request(request)
            return export_daily_reports_csv(qs)
        except Exception as exc:
            from django.contrib import messages
            from django.shortcuts import redirect
            messages.error(request, f"CSV export failed: {exc}")
            return redirect("admin:core_dailyreport_changelist")

    def export_filtered_xlsx_view(self, request):
        try:
            qs = self._filtered_queryset_from_request(request)
            return export_daily_reports_xlsx(qs)
        except Exception as exc:
            from django.contrib import messages
            from django.shortcuts import redirect
            messages.error(request, f"Excel export failed: {exc}")
            return redirect("admin:core_dailyreport_changelist")

    def save_model(self, request, obj, form, change):
        """
        Overrides the admin panel save process. 
        If an admin checks or unchecks a box, it fires our workflow synchronization function.
        """
        # 1. Save the main DailyReport model changes first
        super().save_model(request, obj, form, change)
        
        # 2. Automatically sync sub-models using the logged-in Administrator account
        sync_workflow_records(obj, request.user)
    
    


# --- Standalone Sub-Model Registration ---
# This registers them as independent searchable menus on the Admin homepage
# @admin.register(PDFRecord)
# class PDFRecordAdmin(admin.ModelAdmin):
#     list_display = ('daily_report', 'created_by', 'name', 'created_at')
#     search_fields = ('name', 'daily_report__volume_num')

@admin.register(PDFRecord)
class PDFRecordAdmin(PremiumAdmin):
    # 1. Swap 'daily_report' with our new custom method 'link_to_daily_report'
    list_display = ('link_to_daily_report', 'name', 'created_at')
    search_fields = ('name', 'daily_report__year', 'created_by__username')
    search_help_text = "Search by employee, year, or username"
    list_filter = (
        RelatedReportYearDropdownFilter,
        ('daily_report__location', ChoicesDropdownFilter),
        ('created_by', RelatedDropdownFilter),
        ('created_at', RangeDateFilter),
    )

    def link_to_daily_report(self, obj):
        """
        Generates a secure html anchor tag linking directly 
        to the parent DailyReport change form.
        """
        if obj.daily_report:
            # Reverse lookup the admin URL path dynamically
            # Format: admin:<app_label>_<model_name>_change
            url = reverse('admin:core_dailyreport_change', args=[obj.daily_report.id])
            return format_html('<a href="{}" style="font-weight: bold; color: #15803d;">{}</a>', url, obj.daily_report)
        return "-"

    # 2. Give the column a clean display header and allow it to sort correctly
    link_to_daily_report.short_description = "Daily Report"
    link_to_daily_report.admin_order_field = "daily_report"

@admin.register(IndexingRecord)
class IndexingRecordAdmin(PremiumAdmin):
    list_display = ('link_to_daily_report', 'created_by', 'name', 'created_at')
    search_fields = ('name', 'daily_report__year', 'created_by__username')
    search_help_text = "Search by employee, year, or username"
    list_filter = (
        RelatedReportYearDropdownFilter,
        ('daily_report__location', ChoicesDropdownFilter),
        ('created_by', RelatedDropdownFilter),
        ('created_at', RangeDateFilter),
    )

    def link_to_daily_report(self, obj):
        """
        Generates a secure html anchor tag linking directly 
        to the parent DailyReport change form.
        """
        if obj.daily_report:
            # Reverse lookup the admin URL path dynamically
            # Format: admin:<app_label>_<model_name>_change
            url = reverse('admin:core_dailyreport_change', args=[obj.daily_report.id])
            return format_html('<a href="{}" style="font-weight: bold; color: #15803d;">{}</a>', url, obj.daily_report)
        return "-"

    # 2. Give the column a clean display header and allow it to sort correctly
    link_to_daily_report.short_description = "Daily Report"
    link_to_daily_report.admin_order_field = "daily_report"

@admin.register(UploadingRecord)
class UploadingRecordAdmin(PremiumAdmin):
    list_display = ('link_to_daily_report', 'created_by', 'name', 'created_at')
    search_fields = ('name', 'daily_report__year', 'created_by__username')
    search_help_text = "Search by employee, year, or username"
    list_filter = (
        RelatedReportYearDropdownFilter,
        ('daily_report__location', ChoicesDropdownFilter),
        ('created_by', RelatedDropdownFilter),
        ('created_at', RangeDateFilter),
    )

    def link_to_daily_report(self, obj):
        """
        Generates a secure html anchor tag linking directly 
        to the parent DailyReport change form.
        """
        if obj.daily_report:
            # Reverse lookup the admin URL path dynamically
            # Format: admin:<app_label>_<model_name>_change
            url = reverse('admin:core_dailyreport_change', args=[obj.daily_report.id])
            return format_html('<a href="{}" style="font-weight: bold; color: #15803d;">{}</a>', url, obj.daily_report)
        return "-"

    # 2. Give the column a clean display header and allow it to sort correctly
    link_to_daily_report.short_description = "Daily Report"
    link_to_daily_report.admin_order_field = "daily_report"

@admin.register(QCRecord)
class QCRecordAdmin(PremiumAdmin):
    list_display = ('link_to_daily_report', 'created_by', 'name', 'created_at')
    search_fields = ('name', 'daily_report__year', 'created_by__username')
    search_help_text = "Search by employee, year, or username"
    list_filter = (
        RelatedReportYearDropdownFilter,
        ('daily_report__location', ChoicesDropdownFilter),
        ('created_by', RelatedDropdownFilter),
        ('created_at', RangeDateFilter),
    )

    def link_to_daily_report(self, obj):
        """
        Generates a secure html anchor tag linking directly 
        to the parent DailyReport change form.
        """
        if obj.daily_report:
            # Reverse lookup the admin URL path dynamically
            # Format: admin:<app_label>_<model_name>_change
            url = reverse('admin:core_dailyreport_change', args=[obj.daily_report.id])
            return format_html('<a href="{}" style="font-weight: bold; color: #15803d;">{}</a>', url, obj.daily_report)
        return "-"

    # 2. Give the column a clean display header and allow it to sort correctly
    link_to_daily_report.short_description = "Daily Report"
    link_to_daily_report.admin_order_field = "daily_report"

@admin.register(MetadataRecord)
class MetadataRecordAdmin(PremiumAdmin):
    list_display = ('link_to_daily_report', 'created_by', 'name', 'created_at')
    search_fields =('name', 'daily_report__year', 'created_by__username')
    search_help_text = "Search by employee, year, or username"
    list_filter = (
        RelatedReportYearDropdownFilter,
        ('daily_report__location', ChoicesDropdownFilter),
        ('created_by', RelatedDropdownFilter),
        ('created_at', RangeDateFilter),
    )

    def link_to_daily_report(self, obj):
        """
        Generates a secure html anchor tag linking directly 
        to the parent DailyReport change form.
        """
        if obj.daily_report:
            # Reverse lookup the admin URL path dynamically
            # Format: admin:<app_label>_<model_name>_change
            url = reverse('admin:core_dailyreport_change', args=[obj.daily_report.id])
            return format_html('<a href="{}" style="font-weight: bold; color: #15803d;">{}</a>', url, obj.daily_report)
        return "-"

    # 2. Give the column a clean display header and allow it to sort correctly
    link_to_daily_report.short_description = "Daily Report"
    link_to_daily_report.admin_order_field = "daily_report"